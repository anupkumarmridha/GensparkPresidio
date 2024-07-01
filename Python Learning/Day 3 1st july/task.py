# 1) Create a application that will take the Employee details(Name, DOB, Phone and E-Mail) from console, validate it and calculate age(Age should not taken from user)
#    The application should show menu to store the same in file. Option for saving should be text/excel/pdf. 
#    Optional implementation - > Bulk read and store in list from Excel file

import openpyxl
from fpdf import FPDF
import PyPDF2
from datetime import datetime
import re

class Employee:
    emp_id = 1

    def __init__(self, name, dob, phone, email):
        self.id = Employee.emp_id
        Employee.emp_id += 1
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email
        self.age = self.calculate_age()

    def __str__(self):
        return f"ID: {self.id}, Name: {self.name}, DOB: {self.dob}, Phone: {self.phone}, Email: {self.email}, Age: {self.age}"

    def calculate_age(self):
        dob = datetime.strptime(self.dob, "%d-%m-%Y")
        today = datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age

    def validate(self):
        if not re.match(r"[^@]+@[^@]+\.[^@]+", self.email):
            return False
        if not re.match(r"^[6-9]\d{9}$", self.phone):
            return False
        return True

    def get_details(self):
        return self.name, self.dob, self.phone, self.email, self.age

class EmployeeList:
    def __init__(self):
        self.employees = []

    def add_employee(self, emp):
        self.employees.append(emp)

    def remove_employee(self, emp_id):
        self.employees = [emp for emp in self.employees if emp.id != emp_id]

    def display_employees(self):
        if not self.employees:
            print("No employees")
            return
        for emp in self.employees:
            print(emp)

    def save_to_file(self, file_name):
        with open(file_name, "w") as f:
            for emp in self.employees:
                f.write(str(emp.get_details()) + "\n")

    def save_to_excel(self, file_name):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "DOB", "Phone", "Email", "Age"])
        for emp in self.employees:
            sheet.append(emp.get_details())
        wb.save(file_name)

    def save_to_pdf(self, file_name):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        pdf.cell(200, 10, txt="Employee Details", ln=True, align="C")
        pdf.cell(200, 10, txt="", ln=True, align="C")
        pdf.cell(40, 10, txt="Name", border=1)
        pdf.cell(30, 10, txt="DOB", border=1)
        pdf.cell(30, 10, txt="Phone", border=1)
        pdf.cell(70, 10, txt="Email", border=1)
        pdf.cell(20, 10, txt="Age", border=1)
        pdf.cell(200, 10, txt="", ln=True, align="C")
        for emp in self.employees:
            pdf.cell(40, 10, txt=emp.name, border=1)
            pdf.cell(30, 10, txt=emp.dob, border=1)
            pdf.cell(30, 10, txt=emp.phone, border=1)
            pdf.cell(70, 10, txt=emp.email, border=1)
            pdf.cell(20, 10, txt=str(emp.age), border=1)
            pdf.cell(200, 10, txt="", ln=True, align="C")
        pdf.output(file_name)

    def read_from_text(self, file_name):
        with open(file_name, "r") as f:
            for line in f:
                emp_data = eval(line.strip())
                emp = Employee(*emp_data[:-1])
                self.add_employee(emp)

    def read_from_pdf(self, file_name):
        with open(file_name, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            num_pages = len(reader.pages)
            for i in range(num_pages):
                page = reader.pages[i]
                text = page.extract_text()
                lines = text.split("\n")
                print(lines[2:])
                for line in lines[2:]:
                    emp_data = line.split(" ")
                    if len(emp_data) == 5:
                        name, dob, phone, email, age = emp_data
                        emp = Employee(name.strip(), dob.strip(), phone.strip(), email.strip())
                        self.add_employee(emp)

    def read_from_excel(self, file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            emp = Employee(*row[:-1])
            self.add_employee(emp)

def main():
    emp_list = EmployeeList()
    while True:
        print("1. Add Employee")
        print("2. Display Employees")
        print("3. Remove Employee")
        print("4. Save to text file")
        print("5. Save to excel file")
        print("6. Save to pdf file")
        print("7. Read from text file")
        print("8. Read from excel file")
        print("9. Read from pdf file")
        print("10. Exit")
        
        choice = int(input("Enter your choice: "))
        if choice == 1:
            name = input("Enter name: ")
            dob = input("Enter DOB (dd-mm-yyyy): ")
            phone = input("Enter phone: ")
            email = input("Enter email: ")
            emp = Employee(name, dob, phone, email)
            if emp.validate():
                emp_list.add_employee(emp)
                print("Employee added successfully")
            else:
                print("Invalid data")
        elif choice == 2:
            emp_list.display_employees()
        elif choice == 3:
            emp_id = int(input("Enter Employee ID to remove: "))
            emp_list.remove_employee(emp_id)
            print(f"Employee with ID {emp_id} removed successfully")
        elif choice == 4:
            file_name = input("Enter file name: ")
            emp_list.save_to_file(file_name)
        elif choice == 5:
            file_name = input("Enter file name: ")
            emp_list.save_to_excel(file_name)
        elif choice == 6:
            file_name = input("Enter file name: ")
            emp_list.save_to_pdf(file_name)
        elif choice == 7:
            file_name = input("Enter file name: ")
            emp_list.read_from_text(file_name)
        elif choice == 8:
            file_name = input("Enter file name: ")
            emp_list.read_from_excel(file_name)
        elif choice == 9:
            file_name = input("Enter file name: ")
            emp_list.read_from_pdf(file_name)
        elif choice == 10:
            break
        else:
            print("Invalid choice")

if __name__ == "__main__":
    main()
